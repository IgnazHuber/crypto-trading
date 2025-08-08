# analysis/tradingmodule_250806i.py

import os, sys, json, re, logging, tkinter as tk
import pandas as pd, numpy as np, matplotlib.pyplot as plt
from tkinter import filedialog

from .mod_data import load_data
from .mod_utils import add_columns_from_result
from .mod_strategy import prepare_regimes, generate_signals
from .mod_trades import run_meta_strategy_with_indicators
from .portfolio_metrics import calculate_portfolio_metrics
from .mod_excel import export_to_excel
from .mod_markdown import export_to_markdown
from .mod_csv_html import export_to_csv_and_html
from .mod_png import export_png_equity
from .indicator_descriptions import INDICATOR_INFO
from openpyxl import load_workbook

SKRIPT_ID = "tradingmodule_250806i"
RESULTS_DIR = "results"
os.makedirs(RESULTS_DIR, exist_ok=True)

DEFAULT_PARAMS = {
    "min_signals": 1,
    "ema_length": 20,
    "sma_length": 50,
    "rsi_length": 14,
    "macd_fast": 12,
    "macd_slow": 26,
    "macd_signal": 9,
    "mfi_length": 14,
    "atr_length": 14,
    "bb_length": 20,
    "bb_std": 2.0
}

# --- Parameter laden ---
PARAMS_PATH = os.path.join(RESULTS_DIR, "best_params.json")
global_params, param_origin = {}, {}
if os.path.exists(PARAMS_PATH):
    try:
        with open(PARAMS_PATH, "r") as f:
            global_params = json.load(f)
        print(f"[INFO] Optimierte Parameter aus {PARAMS_PATH} geladen.")
    except Exception as e:
        print(f"[WARN] Fehler beim Laden von best_params.json ({e}), verwende nur Defaultparameter.")
else:
    print("[WARN] Keine best_params.json gefunden, verwende nur Defaultparameter.")

for k, dv in DEFAULT_PARAMS.items():
    if k not in global_params:
        global_params[k] = dv
        param_origin[k] = "default"
    else:
        param_origin[k] = "optimized"

log_path = os.path.join(RESULTS_DIR, f"log_{SKRIPT_ID}.txt")
logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(message)s",
                    handlers=[logging.FileHandler(log_path, mode='w', encoding='utf-8'),
                              logging.StreamHandler()])
logger = logging.getLogger(SKRIPT_ID)


def safe_tab_name(name):
    return re.sub(r'[^A-Za-z0-9_]', '_', name)[:25]


def select_parquet_files():
    root = tk.Tk(); root.withdraw()
    file_paths = filedialog.askopenfilenames(
        title="Wähle eine oder mehrere Parquet-Dateien",
        initialdir=r"D:\Projekte\crypto_trading\crypto_trading\data\raw",
        filetypes=[("Parquet Dateien", "*.parquet")])
    return list(file_paths)


def run_backtest_for_file(parquet_path, trade_every=1, params=None):
    params = params or {}
    asset = os.path.splitext(os.path.basename(parquet_path))[0].split("_")[0]
    df = load_data(parquet_path)
    regimes = prepare_regimes(df)
    add_columns_from_result(df, regimes)
    signals = generate_signals(df, regimes, params)
    df['meta_long'] = signals['meta_long']
    trades = run_meta_strategy_with_indicators(df, 'meta_long', params,
                                               trade_every=trade_every, asset=asset,
                                               strategy_name=f"Meta_Score")
    return asset, trades, df


def build_portfolio_equity(equity_dict):
    merged = pd.DataFrame()
    for file_label, df in equity_dict.items():
        df2 = df.rename(columns={"Equity": file_label})
        merged = pd.merge(merged, df2[["Time", file_label]],
                          how="outer", on="Time") if not merged.empty else df2[["Time", file_label]]
    merged = merged.sort_values("Time").fillna(method="ffill").fillna(method="bfill")
    merged["PortfolioEquity"] = merged.drop(columns="Time").sum(axis=1)
    return merged[["Time", "PortfolioEquity"]]


def export_indicator_table_excel(path, params):
    wb = load_workbook(path)
    ws = wb.create_sheet("Indicators_Used")
    ws.append(["Indikator", "Parameter", "Marktumfeld", "Beschreibung"])
    for k, v in params.items():
        info = INDICATOR_INFO.get(k, {})
        bullets = "\n".join(info.get("bullets", []))
        ws.append([info.get("title", k), v, info.get("context", ""), bullets])
    wb.save(path)


def main():
    parquet_files = select_parquet_files()
    if not parquet_files:
        logger.warning("Keine Dateien ausgewählt."); return

    all_trades_list, equity_dict, all_trades_dict, summary_records = [], {}, {}, []
    for path in parquet_files:
        asset, trades, df = run_backtest_for_file(path, params=global_params)
        equity = 10000 + (trades["pnl_abs"].cumsum() if "pnl_abs" in trades else 0)
        if len(equity) > 0:
            equity.iloc[0] = 10000
        timestamps = trades["exit_time"] if "exit_time" in trades else df["timestamp"]
        equity_dict[os.path.basename(path)] = pd.DataFrame({"Time": timestamps, "Equity": equity, "Drawdown": equity - equity.cummax()})
        export_png_equity(asset, equity_dict[os.path.basename(path)])
        all_trades_dict[os.path.basename(path)] = trades
        summary_records.append({"Asset": asset, "Parquet_File": os.path.basename(path),
                                "End_Capital": equity.iloc[-1], "Trades": len(trades)})
        if not trades.empty:
            trades["Asset"] = asset
            all_trades_list.append((path, trades))

    summary_df = pd.DataFrame(summary_records)
    metrics_df = calculate_portfolio_metrics({lbl: df.set_index("Time")["Equity"] for lbl, df in equity_dict.items()})
    summary_df = pd.merge(summary_df, metrics_df, on="Asset", how="left")
    all_trades_df = pd.concat([t for _, t in all_trades_list], ignore_index=True) if all_trades_list else pd.DataFrame()

    # Standard-Export (mit Diagrammen)
    excel_path = os.path.join(RESULTS_DIR, "trades_summary_all.xlsx")
    export_to_excel(summary_df, all_trades_df, equity_dict)
    export_indicator_table_excel(excel_path, global_params)

    export_to_markdown(summary_df, all_trades_dict)
    export_to_csv_and_html(summary_df, all_trades_df,
                           params=global_params, indicator_info=INDICATOR_INFO,
                           trades_html="results/all_trades.html")

    logger.info("Backtest abgeschlossen.")
