# analysis/tradingmodule_250806k.py

# analysis/tradingmodule_250806_final.py

import os, sys, json, re, logging, tkinter as tk
from tkinter import filedialog
import pandas as pd, numpy as np
from openpyxl import load_workbook

from .mod_data import load_data
from .mod_utils import add_columns_from_result
from .mod_strategy import prepare_regimes, generate_signals
from .mod_trades import run_meta_strategy_with_indicators
from .portfolio_metrics import calculate_portfolio_metrics
from .mod_excel import export_to_excel
from .mod_markdown import export_to_markdown, export_params_to_markdown
from .mod_csv_html import export_to_csv_and_html
from .mod_png import export_png_equity
from .indicator_descriptions import INDICATOR_INFO

SKRIPT_ID = "tradingmodule_250806_final"
RESULTS_DIR = r"D:\Projekte\crypto_trading\results"
LAST_DIR_FILE = os.path.join(RESULTS_DIR, "last_dir.txt")
os.makedirs(RESULTS_DIR, exist_ok=True)

# --- Parameter laden (immer aus fester Datei) ---
PARAMS_PATH = r"D:\Projekte\crypto_trading\results\best_params.json"
if not os.path.exists(PARAMS_PATH):
    raise FileNotFoundError(f"Best-Parameter-Datei fehlt: {PARAMS_PATH}")
with open(PARAMS_PATH, "r") as f:
    global_params = json.load(f)
print(f"[INFO] Optimierte Parameter aus {PARAMS_PATH} geladen.")

log_path = os.path.join(RESULTS_DIR, f"log_{SKRIPT_ID}.txt")
logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(message)s",
                    handlers=[logging.FileHandler(log_path, mode='w', encoding='utf-8'),
                              logging.StreamHandler()])
logger = logging.getLogger(SKRIPT_ID)


def safe_tab_name(name):
    return re.sub(r'[^A-Za-z0-9_]', '_', name)[:25]


def select_parquet_files():
    start_dir = r"D:\Projekte\crypto_trading\crypto_trading\data\raw"
    if os.path.exists(LAST_DIR_FILE):
        with open(LAST_DIR_FILE, "r") as f:
            start_dir = f.read().strip()
    root = tk.Tk(); root.withdraw()
    file_paths = filedialog.askopenfilenames(
        title="Wähle eine oder mehrere Parquet-Dateien",
        initialdir=start_dir,
        filetypes=[("Parquet Dateien", "*.parquet")])
    if file_paths:
        with open(LAST_DIR_FILE, "w") as f:
            f.write(os.path.dirname(file_paths[0]))
    return list(file_paths)


def run_backtest_for_file(parquet_path, trade_every=1, params=None):
    params = params or {}
    asset = os.path.splitext(os.path.basename(parquet_path))[0].split("_")[0]
    logger.info(f"Verarbeite: {asset} mit Parametern: {params}")
    df = load_data(parquet_path)
    regimes = prepare_regimes(df)
    add_columns_from_result(df, regimes)
    signals = generate_signals(df, regimes, params)
    df['meta_long'] = signals['meta_long']
    trades = run_meta_strategy_with_indicators(df, 'meta_long', params,
                                               trade_every=trade_every, asset=asset,
                                               strategy_name="Meta_Score")
    return asset, trades, df


def build_portfolio_equity(equity_dict):
    merged = pd.DataFrame()
    for file_label, df in equity_dict.items():
        df2 = df.rename(columns={"Equity": file_label})
        merged = pd.merge(merged, df2[["Time", file_label]],
                          how="outer", on="Time") if not merged.empty else df2[["Time", file_label]]
    merged = merged.sort_values("Time").fillna(method="ffill").fillna(method="bfill")
    merged["PortfolioEquity"] = merged.drop(columns="Time").sum(axis=1)
    return merged[["Time", "PortfolioEquity"]]


def export_indicator_table_excel(path, params):
    wb = load_workbook(path)
    ws = wb.create_sheet("Indicators_Used")
    ws.append(["Indikator", "Parameter", "Marktumfeld", "Beschreibung"])
    for k, v in params.items():
        info = INDICATOR_INFO.get(k, {})
        bullets = "\n".join(info.get("bullets", []))
        ws.append([info.get("title", k), v, info.get("context", ""), bullets])
    wb.save(path)


def main():
    parquet_files = select_parquet_files()
    if not parquet_files:
        logger.warning("Keine Dateien ausgewählt."); return

    all_trades_list, equity_dict, all_trades_dict, summary_records = [], {}, {}, []
    for path in parquet_files:
        asset, trades, df = run_backtest_for_file(path, params=global_params)
        equity = 10000 + (trades["pnl_abs"].cumsum() if "pnl_abs" in trades else 0)
        if len(equity) > 0:
            equity.iloc[0] = 10000
        timestamps = trades["exit_time"] if "exit_time" in trades else df["timestamp"]
        eq_df = pd.DataFrame({"Time": timestamps, "Equity": equity, "Drawdown": equity - equity.cummax()})
        equity_dict[os.path.basename(path)] = eq_df
        export_png_equity(asset, eq_df)
        logger.info(f"[PNG] Equity/Drawdown gespeichert: results/equity_{asset}.png")
        all_trades_dict[os.path.basename(path)] = trades
        summary_records.append({"Asset": asset, "Parquet_File": os.path.basename(path),
                                "End_Capital": equity.iloc[-1], "Trades": len(trades)})
        if not trades.empty:
            trades["Asset"] = asset
            all_trades_list.append((path, trades))

    summary_df = pd.DataFrame(summary_records)
    metrics_df = calculate_portfolio_metrics({lbl: df.set_index("Time")["Equity"] for lbl, df in equity_dict.items()})
    summary_df = pd.merge(summary_df, metrics_df, on="Asset", how="left")
    all_trades_df = pd.concat([t for _, t in all_trades_list], ignore_index=True) if all_trades_list else pd.DataFrame()

    excel_path = os.path.join(RESULTS_DIR, "trades_summary_all.xlsx")
    export_to_excel(summary_df, all_trades_df, equity_dict)  # enthält Diagramme
    export_indicator_table_excel(excel_path, global_params)
    logger.info(f"[Excel] Export abgeschlossen: {excel_path}")

    export_params_to_markdown(global_params, {}, path=os.path.join(RESULTS_DIR, "used_params.md"))
    logger.info("[Markdown/TXT] Export abgeschlossen.")

    export_to_csv_and_html(summary_df, all_trades_df,
                           params=global_params, indicator_info=INDICATOR_INFO,
                           trades_html=os.path.join(RESULTS_DIR, "all_trades.html"))
    logger.info("[CSV/HTML] Export abgeschlossen.")
    logger.info("Backtest abgeschlossen.")
